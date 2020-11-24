import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Protection

filename="MIU_Key_Highlights_2.xlsx"
sheetname = 'Sector wise Rating'
diag1 = "A1"
diag2 = "I7"

# used for unlocking single or multiple cells
# once unlocked cells remains unlocked
# works for rectangular areas(single cell,line of cells or rectangel of cells)
def unblock_cell(filename,sheetname,diag1,diag2):
    wb = load_workbook(filename)
    ws=wb[str(sheetname)]
    ws.protection.sheet = True
    ws.protection.enable()
    cell=ws[str(diag1):str(diag2)]
    for c_tuple in cell:
        for cl in c_tuple:
            cl.protection = Protection(locked=False)
    wb.save(filename=str(filename))
    
